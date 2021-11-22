# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import os
from openpyxl import Workbook, load_workbook


# 无关紧要
class PapaPipeline(object):
    def process_item(self, item, spider):
        return item
    pass


# # 定义的
# class ggkk(object):
class ggkk(object):  # 设置工序一
    if os.path.exists("job.xlsx") != True:  # 使用函数exists()对文件存在与否进行判断，存在为True,不存在为False.
        wb = Workbook()
        ws = wb.active
        ws.append(['Dig','Img','en_context','Location','Format'])
    else :
        wb = load_workbook('job.xlsx') # 打开一个已有的workbook：
        ws= wb.active



    def process_item(self, item, spider):  # 工序具体内容
        Img = item["Img"]
        Title = item["Title"]
        Alter = item["Alter"]
        Date = item["Date"]
        Creator = item["Creator"]
        Location = item["Location"]
        Format = item["Format"]
        Dig = item["Dig"]
        en_context=Title+"\n\n"+Alter+"\n\n"+Date+"\n\n"+Creator+"\n\n"
        line = [Dig,Img,en_context,Location,Format]  # 把数据每一行整理出来
        self.ws.append(line)  # 将数据一行的形式添加到xlsx中
        self.wb.save('job.xlsx')  # 保存xlsx文件
        return item



    # def __init__(self):
    #     # 加载创建文件
    #     self.file = open("jkl.csv", "a+")
    #
    # def process_item(self, item, spdier):
    #     # 是否为空
    #     if os.path.getsize("jkl.csv"):
    #         # 开始写文件
    #         self.write_content(item)
    #     else:
    #         self.file.write("英文图片内容说明,图片说明,英文关键字,中文关键字\n")  # 第一行
    #     self.file.flush()  # 刷新
    #
    #     return item
    #
    # def write_content(self, item):
    #     # 提取内容
    #     Img = item["Img"]
    #     Title = item["Title"]
    #     Alter = item["Alter"]
    #     Date = item["Date"]
    #     Creator = item["Creator"]
    #     Location = item["Location"]
    #     Format = item["Format"]
    #     Dig = item["Dig"]
    #
    #     # 用self.file.write（）写入
    #     self.file.write('{},{},{},{},{},{},{},{}'.format(Img,Title,Alter,Date,Creator,Location,Format,Dig) + '\n')

